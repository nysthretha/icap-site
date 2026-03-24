from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from holidays import get_holidays_for_month
from models import get_selections_for_month, get_all_doctors


def generate_excel(year, month):
    """Generate an Excel file for the finalized schedule."""
    wb = Workbook()
    ws = wb.active

    turkish_months = [
        "", "Ocak", "Subat", "Mart", "Nisan", "Mayis", "Haziran",
        "Temmuz", "Agustos", "Eylul", "Ekim", "Kasim", "Aralik"
    ]
    ws.title = f"{turkish_months[month]} {year} Nobet"

    # Get data
    days = get_holidays_for_month(year, month)
    selections = get_selections_for_month(year, month)
    doctors = get_all_doctors()

    # Unique specialties sorted
    specialties = sorted(set(d["specialty"] for d in doctors))

    # Build selection lookup: {(date, specialty): doctor_name}
    sel_lookup = {}
    for s in selections:
        sel_lookup[(s["date"], s["specialty"])] = s["full_name"]

    # Styles
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    grey_fill = PatternFill("solid", fgColor="D9D9D9")
    header_fill = PatternFill("solid", fgColor="B4C6E7")
    header_font = Font(bold=True, size=11)
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(specialties) + 3)
    title_cell = ws.cell(row=1, column=1,
                         value=f"{turkish_months[month]} {year} - Nobet Cizelgesi")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align

    # Header row
    header_row = 3
    headers = ["Tarih", "Gun", "Saat"] + specialties
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Data rows
    for row_idx, day_info in enumerate(days, start=header_row + 1):
        date_str = day_info["date"]
        is_grey = day_info["type"] in ("holiday", "weekend")
        fill = grey_fill if is_grey else white_fill

        # Date column
        cell = ws.cell(row=row_idx, column=1, value=date_str)
        cell.fill = fill
        cell.border = thin_border
        cell.font = cell_font
        cell.alignment = center_align

        # Day name column
        cell = ws.cell(row=row_idx, column=2, value=day_info["day_name"])
        cell.fill = fill
        cell.border = thin_border
        cell.font = cell_font
        cell.alignment = center_align

        # Duty hours column
        cell = ws.cell(row=row_idx, column=3, value=f"{day_info['duty_hours']} saat")
        cell.fill = fill
        cell.border = thin_border
        cell.font = cell_font
        cell.alignment = center_align

        # Specialty columns
        for col_idx, spec in enumerate(specialties, start=4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = thin_border
            cell.font = cell_font
            cell.alignment = center_align
            assigned = sel_lookup.get((date_str, spec))
            if assigned:
                cell.value = assigned

    # Total hours row
    totals_row = header_row + 1 + len(days)
    totals_by_spec = {spec: 0 for spec in specialties}
    for s in selections:
        totals_by_spec[s["specialty"]] = totals_by_spec.get(s["specialty"], 0) + s["duty_hours"]

    total_font = Font(bold=True, size=11)
    total_fill = PatternFill("solid", fgColor="B4C6E7")

    cell = ws.cell(row=totals_row, column=1, value="Toplam Saat")
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = center_align
    for col_idx in [2, 3]:
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.fill = total_fill
        cell.border = thin_border
    for col_idx, spec in enumerate(specialties, start=4):
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = center_align
        total = totals_by_spec.get(spec, 0)
        cell.value = f"{total} saat" if total > 0 else "-"

    # Auto-adjust column widths
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = max(max_length + 3, 12)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
